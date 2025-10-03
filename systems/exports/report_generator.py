"""
EdTech RADAR - Advanced Report Generator
=======================================

Generate comprehensive reports in multiple formats (PDF, Word, Excel, Markdown).
Supports automated analysis, executive summaries, and detailed company profiles.
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import json
import logging
from pathlib import Path
import io
import base64

# Optional imports for enhanced functionality
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available. PDF generation will be limited.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL not available. Excel generation will be limited.")

from systems.models.edtech_schemas import CompanyProfile, MarketOpportunity, AnalysisScore
from systems.analysis.scoring_engine import EdTechScoringEngine
from systems.data.data_manager import EdTechDataManager


class EdTechReportGenerator:
    """Comprehensive report generation for EdTech market intelligence"""

    def __init__(self, data_manager: EdTechDataManager, scoring_engine: EdTechScoringEngine):
        self.data_manager = data_manager
        self.scoring_engine = scoring_engine
        self.logger = logging.getLogger(__name__)

        # Report templates and styling
        self.report_styles = {
            'title': {'font_size': 24, 'font_weight': 'bold', 'color': '#1f77b4'},
            'heading1': {'font_size': 18, 'font_weight': 'bold', 'color': '#2c3e50'},
            'heading2': {'font_size': 14, 'font_weight': 'bold', 'color': '#34495e'},
            'body': {'font_size': 11, 'color': '#2c3e50'},
            'highlight': {'background_color': '#f8f9fa', 'border': '1px solid #dee2e6'}
        }

    def generate_executive_summary(self, companies: List[CompanyProfile],
                                 scores: List[AnalysisScore],
                                 opportunities: List[MarketOpportunity]) -> Dict[str, Any]:
        """Generate executive summary with key insights and recommendations"""

        if not companies or not scores:
            return {'error': 'No data available for analysis'}

        # Calculate key metrics
        total_companies = len(companies)
        avg_score = np.mean([score.total_score for score in scores])
        top_quartile_threshold = np.percentile([score.total_score for score in scores], 75)
        strong_buy_count = sum(1 for score in scores if score.recommendation == 'STRONG_BUY')
        buy_count = sum(1 for score in scores if score.recommendation == 'BUY')

        # Funding analysis
        total_funding = sum(company.funding.total_raised or 0 for company in companies)
        avg_funding = total_funding / total_companies if total_companies > 0 else 0

        # Category analysis
        category_performance = self._analyze_category_performance(companies, scores)
        top_category = max(category_performance.items(), key=lambda x: x[1]['avg_score']) if category_performance else None

        # Market opportunities
        total_market_size = sum(opp.market_size or 0 for opp in opportunities)
        avg_growth_rate = np.mean([opp.growth_rate for opp in opportunities if opp.growth_rate])

        # Generate insights and recommendations
        insights = self._generate_key_insights(companies, scores, opportunities)
        recommendations = self._generate_strategic_recommendations(companies, scores)

        executive_summary = {
            'generated_at': datetime.now().isoformat(),
            'analysis_period': f"{datetime.now().strftime('%B %Y')}",
            'key_metrics': {
                'total_companies_analyzed': total_companies,
                'average_portfolio_score': round(avg_score, 1),
                'top_quartile_threshold': round(top_quartile_threshold, 1),
                'investment_recommendations': {
                    'strong_buy': strong_buy_count,
                    'buy': buy_count,
                    'hold': sum(1 for score in scores if score.recommendation == 'HOLD'),
                    'sell': sum(1 for score in scores if score.recommendation == 'SELL')
                },
                'total_funding_analyzed': f"${total_funding:,.0f}",
                'average_funding_per_company': f"${avg_funding:,.0f}"
            },
            'market_overview': {
                'total_addressable_market': f"${total_market_size:,.0f}",
                'average_market_growth_rate': f"{avg_growth_rate:.1f}%" if not np.isnan(avg_growth_rate) else "N/A",
                'total_opportunities_identified': len(opportunities),
                'top_performing_category': top_category[0] if top_category else "N/A",
                'top_category_avg_score': round(top_category[1]['avg_score'], 1) if top_category else 0
            },
            'key_insights': insights,
            'strategic_recommendations': recommendations,
            'portfolio_composition': self._analyze_portfolio_composition(companies),
            'risk_assessment': self._assess_portfolio_risk(companies, scores)
        }

        return executive_summary

    def generate_detailed_company_report(self, company: CompanyProfile, score: AnalysisScore) -> Dict[str, Any]:
        """Generate detailed analysis report for a specific company"""

        report = {
            'company_overview': {
                'name': company.name,
                'website': company.website,
                'founded': company.founded,
                'description': company.description,
                'headquarters': company.geographic_presence.headquarters,
                'employee_count': company.metrics.employees_count,
                'categories': [cat.value for cat in company.category] if company.category else [],
                'target_audiences': [aud.value for aud in company.target_audience] if company.target_audience else [],
                'business_models': [bm.value for bm in company.business_model] if company.business_model else []
            },
            'financial_analysis': {
                'total_funding_raised': company.funding.total_raised,
                'latest_round': company.funding.latest_round,
                'latest_round_amount': company.funding.latest_round_amount,
                'funding_stage': company.funding.stage.value if company.funding.stage else None,
                'investors': company.funding.investors,
                'valuation': company.funding.valuation,
                'annual_revenue': company.metrics.annual_revenue,
                'user_base': company.metrics.user_base,
                'active_users': company.metrics.active_users,
                'growth_rate': company.metrics.growth_rate,
                'retention_rate': company.metrics.retention_rate
            },
            'technology_assessment': {
                'frontend_technologies': company.technology_stack.frontend,
                'backend_technologies': company.technology_stack.backend,
                'databases': company.technology_stack.database,
                'ai_ml_technologies': company.technology_stack.ai_ml,
                'cloud_platforms': company.technology_stack.cloud_platform,
                'mobile_technologies': company.technology_stack.mobile,
                'technology_maturity_score': score.technology_score
            },
            'market_position': {
                'primary_markets': company.geographic_presence.primary_markets,
                'expansion_markets': company.geographic_presence.expansion_markets,
                'total_countries': company.geographic_presence.total_countries,
                'key_competitors': company.competitors,
                'strategic_partnerships': company.partnerships,
                'market_share': company.metrics.market_share,
                'competitive_advantages': company.competitive_advantages
            },
            'investment_analysis': {
                'overall_score': score.total_score,
                'investment_grade': score.investment_grade,
                'recommendation': score.recommendation,
                'detailed_scores': {
                    'market_attractiveness': {
                        'market_size': score.market_size_score,
                        'growth_potential': score.growth_potential_score,
                        'competitive_landscape': score.competitive_landscape_score
                    },
                    'company_strength': {
                        'financial_strength': score.financial_strength_score,
                        'technology': score.technology_score,
                        'team': score.team_score,
                        'product': score.product_score
                    },
                    'strategic_fit': {
                        'alignment': score.alignment_score,
                        'synergy_potential': score.synergy_potential_score,
                        'risk_assessment': score.risk_assessment_score
                    }
                }
            },
            'swot_analysis': {
                'strengths': company.competitive_advantages,
                'weaknesses': company.weaknesses,
                'opportunities': company.opportunities,
                'threats': company.threats
            },
            'products_services': [
                {
                    'name': product.name,
                    'description': product.description,
                    'category': product.category.value if product.category else None,
                    'target_audience': [aud.value for aud in product.target_audience],
                    'features': product.features,
                    'pricing': product.pricing,
                    'launch_date': product.launch_date.isoformat() if product.launch_date else None,
                    'status': product.status
                }
                for product in company.products
            ],
            'generated_at': datetime.now().isoformat(),
            'data_confidence': company.confidence_score,
            'data_sources': company.data_sources
        }

        return report

    def generate_market_analysis_report(self, opportunities: List[MarketOpportunity]) -> Dict[str, Any]:
        """Generate comprehensive market analysis report"""

        if not opportunities:
            return {'error': 'No market opportunities available for analysis'}

        # Market sizing analysis
        total_market_size = sum(opp.market_size or 0 for opp in opportunities)
        market_sizes = [opp.market_size for opp in opportunities if opp.market_size]
        avg_market_size = np.mean(market_sizes) if market_sizes else 0
        median_market_size = np.median(market_sizes) if market_sizes else 0

        # Growth analysis
        growth_rates = [opp.growth_rate for opp in opportunities if opp.growth_rate]
        avg_growth = np.mean(growth_rates) if growth_rates else 0
        high_growth_count = sum(1 for rate in growth_rates if rate > 25)

        # Investment analysis
        investment_amounts = [opp.investment_needed for opp in opportunities if opp.investment_needed]
        total_investment_needed = sum(investment_amounts) if investment_amounts else 0
        avg_investment = np.mean(investment_amounts) if investment_amounts else 0

        # ROI analysis
        roi_potentials = [opp.roi_potential for opp in opportunities if opp.roi_potential]
        avg_roi = np.mean(roi_potentials) if roi_potentials else 0
        high_roi_count = sum(1 for roi in roi_potentials if roi > 50)

        # Category analysis
        category_analysis = self._analyze_opportunity_categories(opportunities)

        # Regional analysis
        regional_analysis = self._analyze_regional_opportunities(opportunities)

        report = {
            'executive_summary': {
                'total_opportunities': len(opportunities),
                'total_addressable_market': f"${total_market_size:,.0f}",
                'average_market_size': f"${avg_market_size:,.0f}",
                'median_market_size': f"${median_market_size:,.0f}",
                'average_growth_rate': f"{avg_growth:.1f}%",
                'high_growth_opportunities': high_growth_count,
                'total_investment_required': f"${total_investment_needed:,.0f}",
                'average_investment_per_opportunity': f"${avg_investment:,.0f}",
                'average_roi_potential': f"{avg_roi:.1f}%",
                'high_roi_opportunities': high_roi_count
            },
            'market_sizing': {
                'total_tam': total_market_size,
                'size_distribution': self._calculate_size_distribution(market_sizes),
                'growth_segments': self._analyze_growth_segments(opportunities)
            },
            'category_analysis': category_analysis,
            'regional_analysis': regional_analysis,
            'investment_analysis': {
                'total_capital_required': total_investment_needed,
                'investment_distribution': self._analyze_investment_distribution(opportunities),
                'roi_analysis': self._analyze_roi_distribution(opportunities)
            },
            'risk_assessment': {
                'risk_distribution': self._analyze_risk_distribution(opportunities),
                'high_risk_opportunities': [opp.name for opp in opportunities if (opp.risk_level or 0) > 7],
                'low_risk_opportunities': [opp.name for opp in opportunities if (opp.risk_level or 0) < 4]
            },
            'strategic_recommendations': self._generate_market_recommendations(opportunities),
            'opportunity_ranking': self._rank_opportunities(opportunities),
            'generated_at': datetime.now().isoformat()
        }

        return report

    def export_to_markdown(self, report_data: Dict[str, Any], filename: str) -> str:
        """Export report to Markdown format"""

        markdown_content = f"""# EdTech RADAR - Market Intelligence Report

*Generated on: {datetime.now().strftime('%B %d, %Y')}*

---

## Executive Summary

"""

        if 'key_metrics' in report_data:
            markdown_content += f"""
### Key Metrics
- **Total Companies Analyzed**: {report_data['key_metrics']['total_companies_analyzed']}
- **Average Portfolio Score**: {report_data['key_metrics']['average_portfolio_score']}
- **Total Funding Analyzed**: {report_data['key_metrics']['total_funding_analyzed']}

### Investment Recommendations
- **Strong Buy**: {report_data['key_metrics']['investment_recommendations']['strong_buy']} companies
- **Buy**: {report_data['key_metrics']['investment_recommendations']['buy']} companies
- **Hold**: {report_data['key_metrics']['investment_recommendations']['hold']} companies
- **Sell**: {report_data['key_metrics']['investment_recommendations']['sell']} companies

"""

        if 'market_overview' in report_data:
            markdown_content += f"""
### Market Overview
- **Total Addressable Market**: {report_data['market_overview']['total_addressable_market']}
- **Average Growth Rate**: {report_data['market_overview']['average_market_growth_rate']}
- **Top Performing Category**: {report_data['market_overview']['top_performing_category']}

"""

        if 'key_insights' in report_data:
            markdown_content += """
## Key Insights

"""
            for insight in report_data['key_insights']:
                markdown_content += f"- {insight}\n"

        if 'strategic_recommendations' in report_data:
            markdown_content += """
## Strategic Recommendations

"""
            for recommendation in report_data['strategic_recommendations']:
                markdown_content += f"- {recommendation}\n"

        # Add company-specific sections if this is a detailed company report
        if 'company_overview' in report_data:
            company = report_data['company_overview']
            markdown_content += f"""
## Company Profile: {company['name']}

### Overview
- **Website**: {company['website']}
- **Founded**: {company['founded']}
- **Headquarters**: {company['headquarters']}
- **Employees**: {company['employee_count']}

### Business Model
- **Categories**: {', '.join(company['categories'])}
- **Target Audiences**: {', '.join(company['target_audiences'])}
- **Business Models**: {', '.join(company['business_models'])}

### Financial Highlights
- **Total Funding**: ${report_data['financial_analysis']['total_funding_raised']:,.0f}
- **Latest Round**: {report_data['financial_analysis']['latest_round']}
- **Funding Stage**: {report_data['financial_analysis']['funding_stage']}
- **User Base**: {report_data['financial_analysis']['user_base']:,} users

### Investment Analysis
- **Overall Score**: {report_data['investment_analysis']['overall_score']:.1f}
- **Investment Grade**: {report_data['investment_analysis']['investment_grade']}
- **Recommendation**: {report_data['investment_analysis']['recommendation']}

"""

        # Save to file
        output_path = Path("systems/exports") / f"{filename}.md"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)

        self.logger.info(f"Markdown report exported to: {output_path}")
        return str(output_path)

    def export_to_excel(self, companies: List[CompanyProfile],
                       scores: List[AnalysisScore],
                       opportunities: List[MarketOpportunity],
                       filename: str) -> str:
        """Export comprehensive data to Excel workbook"""

        if not OPENPYXL_AVAILABLE:
            self.logger.error("OpenPyXL not available. Cannot generate Excel reports.")
            return ""

        output_path = Path("systems/exports") / f"{filename}.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Company Overview Sheet
        self._create_companies_sheet(wb, companies, scores)

        # Market Opportunities Sheet
        self._create_opportunities_sheet(wb, opportunities)

        # Analysis Summary Sheet
        self._create_summary_sheet(wb, companies, scores, opportunities)

        # Score Breakdown Sheet
        self._create_scores_sheet(wb, companies, scores)

        # Save workbook
        wb.save(output_path)
        self.logger.info(f"Excel report exported to: {output_path}")
        return str(output_path)

    def export_to_json(self, report_data: Dict[str, Any], filename: str) -> str:
        """Export report data to JSON format"""

        output_path = Path("systems/exports") / f"{filename}.json"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert datetime objects to strings for JSON serialization
        json_data = self._serialize_for_json(report_data)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)

        self.logger.info(f"JSON report exported to: {output_path}")
        return str(output_path)

    # Helper methods for analysis and data processing

    def _analyze_category_performance(self, companies: List[CompanyProfile],
                                    scores: List[AnalysisScore]) -> Dict[str, Dict[str, Any]]:
        """Analyze performance by category"""
        category_data = {}

        for company, score in zip(companies, scores):
            for category in company.category:
                cat_name = category.value
                if cat_name not in category_data:
                    category_data[cat_name] = {'scores': [], 'companies': []}

                category_data[cat_name]['scores'].append(score.total_score)
                category_data[cat_name]['companies'].append(company.name)

        # Calculate statistics
        for cat_name, data in category_data.items():
            scores = data['scores']
            data['avg_score'] = np.mean(scores)
            data['max_score'] = max(scores)
            data['min_score'] = min(scores)
            data['count'] = len(scores)
            data['std_score'] = np.std(scores)

        return category_data

    def _generate_key_insights(self, companies: List[CompanyProfile],
                             scores: List[AnalysisScore],
                             opportunities: List[MarketOpportunity]) -> List[str]:
        """Generate key insights from the analysis"""
        insights = []

        # Portfolio performance insights
        total_scores = [score.total_score for score in scores]
        avg_score = np.mean(total_scores)

        if avg_score > 75:
            insights.append("Portfolio demonstrates strong overall performance with high-quality companies")
        elif avg_score > 60:
            insights.append("Portfolio shows moderate performance with room for optimization")
        else:
            insights.append("Portfolio requires significant improvements to meet investment standards")

        # Funding insights
        well_funded_count = sum(1 for company in companies if (company.funding.total_raised or 0) > 10000000)
        if well_funded_count > len(companies) * 0.5:
            insights.append("Majority of portfolio companies are well-funded, indicating strong investor confidence")

        # Growth insights
        high_growth_count = sum(1 for company in companies if (company.metrics.growth_rate or 0) > 50)
        if high_growth_count > 0:
            insights.append(f"{high_growth_count} companies show exceptional growth rates above 50% YoY")

        # Technology insights
        ai_companies = sum(1 for company in companies if company.technology_stack.ai_ml)
        if ai_companies > len(companies) * 0.3:
            insights.append("Strong AI/ML adoption across portfolio indicates technological advancement")

        # Market opportunity insights
        if opportunities:
            high_growth_opps = sum(1 for opp in opportunities if (opp.growth_rate or 0) > 25)
            if high_growth_opps > 0:
                insights.append(f"{high_growth_opps} market opportunities show high growth potential above 25%")

        return insights

    def _generate_strategic_recommendations(self, companies: List[CompanyProfile],
                                          scores: List[AnalysisScore]) -> List[str]:
        """Generate strategic recommendations"""
        recommendations = []

        # Investment recommendations
        strong_performers = sum(1 for score in scores if score.total_score > 80)
        if strong_performers > 0:
            recommendations.append(f"Prioritize {strong_performers} top-performing companies for increased investment")

        # Portfolio diversification
        categories = set()
        for company in companies:
            categories.update(cat.value for cat in company.category)

        if len(categories) < 5:
            recommendations.append("Consider diversifying portfolio across more EdTech categories")

        # Technology recommendations
        ai_adoption = sum(1 for company in companies if company.technology_stack.ai_ml) / len(companies) * 100
        if ai_adoption < 50:
            recommendations.append("Increase focus on AI/ML-enabled EdTech companies for future growth")

        # Geographic recommendations
        us_companies = sum(1 for company in companies
                          if company.geographic_presence.headquarters == "United States")
        if us_companies < len(companies) * 0.3:
            recommendations.append("Consider increasing exposure to US EdTech market")

        return recommendations

    def _analyze_portfolio_composition(self, companies: List[CompanyProfile]) -> Dict[str, Any]:
        """Analyze portfolio composition"""
        composition = {
            'by_category': {},
            'by_funding_stage': {},
            'by_geography': {},
            'by_business_model': {}
        }

        # Category composition
        for company in companies:
            for category in company.category:
                cat_name = category.value
                composition['by_category'][cat_name] = composition['by_category'].get(cat_name, 0) + 1

        # Funding stage composition
        for company in companies:
            if company.funding.stage:
                stage = company.funding.stage.value
                composition['by_funding_stage'][stage] = composition['by_funding_stage'].get(stage, 0) + 1

        # Geographic composition
        for company in companies:
            if company.geographic_presence.headquarters:
                geo = company.geographic_presence.headquarters
                composition['by_geography'][geo] = composition['by_geography'].get(geo, 0) + 1

        # Business model composition
        for company in companies:
            for model in company.business_model:
                model_name = model.value
                composition['by_business_model'][model_name] = composition['by_business_model'].get(model_name, 0) + 1

        return composition

    def _assess_portfolio_risk(self, companies: List[CompanyProfile],
                             scores: List[AnalysisScore]) -> Dict[str, Any]:
        """Assess overall portfolio risk"""
        risk_scores = [score.risk_assessment_score for score in scores]
        avg_risk = np.mean(risk_scores)

        # Funding concentration risk
        total_funding = sum(company.funding.total_raised or 0 for company in companies)
        funding_amounts = [company.funding.total_raised or 0 for company in companies]
        funding_concentration = max(funding_amounts) / total_funding if total_funding > 0 else 0

        # Geographic concentration risk
        geo_counts = {}
        for company in companies:
            if company.geographic_presence.headquarters:
                geo = company.geographic_presence.headquarters
                geo_counts[geo] = geo_counts.get(geo, 0) + 1

        max_geo_concentration = max(geo_counts.values()) / len(companies) if geo_counts else 0

        risk_level = "Low"
        if avg_risk < 60 or funding_concentration > 0.5 or max_geo_concentration > 0.6:
            risk_level = "High"
        elif avg_risk < 70 or funding_concentration > 0.3 or max_geo_concentration > 0.4:
            risk_level = "Medium"

        return {
            'overall_risk_level': risk_level,
            'average_risk_score': round(avg_risk, 1),
            'funding_concentration': round(funding_concentration * 100, 1),
            'geographic_concentration': round(max_geo_concentration * 100, 1),
            'risk_factors': self._identify_risk_factors(companies, scores)
        }

    def _identify_risk_factors(self, companies: List[CompanyProfile],
                             scores: List[AnalysisScore]) -> List[str]:
        """Identify key risk factors in the portfolio"""
        risk_factors = []

        # Early stage concentration
        early_stage_count = sum(1 for company in companies
                               if company.funding.stage and
                               company.funding.stage.value in ['bootstrap', 'pre_seed', 'seed'])

        if early_stage_count > len(companies) * 0.6:
            risk_factors.append("High concentration of early-stage companies")

        # Low user base companies
        low_users_count = sum(1 for company in companies
                             if (company.metrics.user_base or 0) < 10000)

        if low_users_count > len(companies) * 0.5:
            risk_factors.append("Many companies have limited user traction")

        # Competitive market exposure
        high_competition_count = sum(1 for company in companies
                                   if len(company.competitors) > 10)

        if high_competition_count > len(companies) * 0.4:
            risk_factors.append("Significant exposure to highly competitive markets")

        return risk_factors

    # Additional helper methods for Excel generation and data processing...

    def _create_companies_sheet(self, wb: 'Workbook', companies: List[CompanyProfile],
                              scores: List[AnalysisScore]):
        """Create companies overview sheet in Excel workbook"""
        ws = wb.create_sheet("Companies Overview")

        headers = [
            "Company Name", "Website", "Founded", "Category", "Funding Stage",
            "Total Funding", "Employees", "User Base", "Headquarters",
            "Overall Score", "Investment Grade", "Recommendation"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write data
        for row, (company, score) in enumerate(zip(companies, scores), 2):
            ws.cell(row=row, column=1, value=company.name)
            ws.cell(row=row, column=2, value=company.website)
            ws.cell(row=row, column=3, value=company.founded)
            ws.cell(row=row, column=4, value=", ".join([cat.value for cat in company.category]))
            ws.cell(row=row, column=5, value=company.funding.stage.value if company.funding.stage else "")
            ws.cell(row=row, column=6, value=company.funding.total_raised or 0)
            ws.cell(row=row, column=7, value=company.metrics.employees_count or 0)
            ws.cell(row=row, column=8, value=company.metrics.user_base or 0)
            ws.cell(row=row, column=9, value=company.geographic_presence.headquarters or "")
            ws.cell(row=row, column=10, value=round(score.total_score, 1))
            ws.cell(row=row, column=11, value=score.investment_grade)
            ws.cell(row=row, column=12, value=score.recommendation)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_opportunities_sheet(self, wb: 'Workbook', opportunities: List[MarketOpportunity]):
        """Create market opportunities sheet in Excel workbook"""
        ws = wb.create_sheet("Market Opportunities")

        headers = [
            "Opportunity ID", "Name", "Category", "Market Size", "Growth Rate",
            "Investment Needed", "ROI Potential", "Risk Level", "Time to Market"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write data
        for row, opportunity in enumerate(opportunities, 2):
            ws.cell(row=row, column=1, value=opportunity.id)
            ws.cell(row=row, column=2, value=opportunity.name)
            ws.cell(row=row, column=3, value=opportunity.category.value if opportunity.category else "")
            ws.cell(row=row, column=4, value=opportunity.market_size or 0)
            ws.cell(row=row, column=5, value=opportunity.growth_rate or 0)
            ws.cell(row=row, column=6, value=opportunity.investment_needed or 0)
            ws.cell(row=row, column=7, value=opportunity.roi_potential or 0)
            ws.cell(row=row, column=8, value=opportunity.risk_level or 0)
            ws.cell(row=row, column=9, value=opportunity.time_to_market or 0)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _serialize_for_json(self, data: Any) -> Any:
        """Recursively serialize data for JSON export"""
        if isinstance(data, datetime):
            return data.isoformat()
        elif isinstance(data, dict):
            return {key: self._serialize_for_json(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [self._serialize_for_json(item) for item in data]
        elif hasattr(data, '__dict__'):
            return self._serialize_for_json(data.__dict__)
        else:
            return data